"""
ExportAgent – Subagent 6.

Exportiert Kalkulationsergebnisse in verschiedene Formate:
- Angebots-PDF (Brandstifter-Design)
- Interne Kalkulations-PDF (vollstaendige Kostenaufstellung)
- Excel-Export (fuer manuelle Nachbearbeitung)
- GAEB-Datei (ausgefuelltes LV)
- Schreiners Buero API Export
"""

from __future__ import annotations

import json
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any

from agents.base_agent import AgentMessage, BaseAgent


class ExportAgent(BaseAgent):
    """Exportiert Kalkulationsergebnisse in alle benoetigten Formate."""

    def __init__(self) -> None:
        super().__init__(name="export_agent")
        self._export_verzeichnis = Path("exports")
        self._sb_api_url = ""
        self._sb_api_user = ""
        self._sb_api_pw = ""

    def configure(
        self,
        export_verzeichnis: str = "exports",
        sb_api_url: str = "",
        sb_api_user: str = "",
        sb_api_pw: str = "",
    ) -> None:
        self._export_verzeichnis = Path(export_verzeichnis)
        self._export_verzeichnis.mkdir(parents=True, exist_ok=True)
        self._sb_api_url = sb_api_url
        self._sb_api_user = sb_api_user
        self._sb_api_pw = sb_api_pw

    async def process(self, message: AgentMessage) -> AgentMessage:
        msg_type = message.msg_type
        kalkulation = message.payload.get("kalkulation", {})
        projekt_id = message.projekt_id

        if msg_type == "export_alle":
            return await self._export_alle(message, kalkulation, projekt_id)

        export_map = {
            "export_angebot_pdf": self._export_angebot_pdf,
            "export_intern_pdf": self._export_intern_pdf,
            "export_excel": self._export_excel,
            "export_gaeb": self._export_gaeb,
            "export_schreiners_buero": self._export_schreiners_buero,
        }

        handler = export_map.get(msg_type)
        if handler is None:
            return message.create_error(
                sender=self.name,
                error_msg=f"Unbekannter Export-Typ: {msg_type}",
            )

        result = await handler(kalkulation, projekt_id)
        return message.create_response(sender=self.name, payload=result)

    async def _export_alle(
        self, message: AgentMessage, kalkulation: dict, projekt_id: str
    ) -> AgentMessage:
        ergebnisse: dict[str, Any] = {}
        for name, handler in [
            ("angebot_pdf", self._export_angebot_pdf),
            ("intern_pdf", self._export_intern_pdf),
            ("excel", self._export_excel),
        ]:
            try:
                ergebnisse[name] = await handler(kalkulation, projekt_id)
            except Exception as exc:
                ergebnisse[name] = {"status": "fehler", "message": str(exc)}

        return message.create_response(
            sender=self.name, payload={"exports": ergebnisse}
        )

    # ------------------------------------------------------------------
    # Angebots-PDF (Brandstifter-Design)
    # ------------------------------------------------------------------

    async def _export_angebot_pdf(
        self, kalkulation: dict, projekt_id: str
    ) -> dict[str, Any]:
        """Generiert professionelles Angebots-PDF."""
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
        )
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_LEFT, TA_RIGHT, TA_CENTER

        dateiname = f"Angebot_{projekt_id}_{datetime.now().strftime('%Y%m%d')}.pdf"
        pfad = self._export_verzeichnis / dateiname

        doc = SimpleDocTemplate(
            str(pfad), pagesize=A4,
            leftMargin=20 * mm, rightMargin=20 * mm,
            topMargin=25 * mm, bottomMargin=20 * mm,
        )

        styles = getSampleStyleSheet()
        styles.add(ParagraphStyle(
            "BrandHeader", parent=styles["Heading1"],
            fontSize=18, textColor=colors.HexColor("#1a1a1a"),
            spaceAfter=6 * mm,
        ))
        styles.add(ParagraphStyle(
            "BrandSub", parent=styles["Normal"],
            fontSize=10, textColor=colors.HexColor("#666666"),
        ))
        styles.add(ParagraphStyle(
            "PosText", parent=styles["Normal"], fontSize=9,
        ))
        styles.add(ParagraphStyle(
            "PosRight", parent=styles["Normal"], fontSize=9, alignment=TA_RIGHT,
        ))
        styles.add(ParagraphStyle(
            "TotalLabel", parent=styles["Normal"],
            fontSize=11, fontName="Helvetica-Bold",
        ))
        styles.add(ParagraphStyle(
            "TotalValue", parent=styles["Normal"],
            fontSize=11, fontName="Helvetica-Bold", alignment=TA_RIGHT,
        ))

        zuschlaege = kalkulation.get("zuschlaege", {})
        positionen = kalkulation.get("positionen", [])

        story: list = []

        # Header
        story.append(Paragraph("AMP &amp; Brandstifter GmbH", styles["BrandHeader"]))
        story.append(Paragraph(
            "Schreinerei &amp; Moebelmanufaktur | Ober-Moerlen", styles["BrandSub"]
        ))
        story.append(Spacer(1, 10 * mm))

        # Angebotsdaten
        story.append(Paragraph(
            f"<b>Angebot {projekt_id}</b>", styles["Heading2"]
        ))
        story.append(Paragraph(
            f"Datum: {datetime.now().strftime('%d.%m.%Y')}", styles["Normal"]
        ))
        story.append(Spacer(1, 8 * mm))

        # Positionstabelle
        header_row = ["Pos.", "Bezeichnung", "Menge", "Einheit", "EP (EUR)", "GP (EUR)"]
        table_data = [header_row]

        for pos in positionen:
            ep = zuschlaege.get("angebotspreis_gesamt", 0) / max(len(positionen), 1)
            ep_pro_einheit = ep / max(float(pos.get("menge", 1)), 1)
            gp = ep

            table_data.append([
                Paragraph(str(pos.get("pos_nr", "")), styles["PosText"]),
                Paragraph(str(pos.get("kurztext", "")), styles["PosText"]),
                Paragraph(f"{float(pos.get('menge', 0)):.1f}", styles["PosRight"]),
                Paragraph(str(pos.get("einheit", "STK")), styles["PosText"]),
                Paragraph(f"{ep_pro_einheit:,.2f}", styles["PosRight"]),
                Paragraph(f"{gp:,.2f}", styles["PosRight"]),
            ])

        col_widths = [15 * mm, 70 * mm, 18 * mm, 15 * mm, 25 * mm, 25 * mm]
        table = Table(table_data, colWidths=col_widths)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2d2d2d")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
            ("TOPPADDING", (0, 0), (-1, 0), 8),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cccccc")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f5f5f5")]),
            ("TOPPADDING", (0, 1), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 1), (-1, -1), 4),
        ]))
        story.append(table)
        story.append(Spacer(1, 8 * mm))

        # Gesamtsumme
        angebotspreis = zuschlaege.get("angebotspreis_gesamt", 0)
        netto = angebotspreis
        mwst = netto * 0.19
        brutto = netto + mwst

        summen_data = [
            ["Nettosumme:", f"{netto:,.2f} EUR"],
            ["zzgl. 19% MwSt.:", f"{mwst:,.2f} EUR"],
            ["Bruttosumme:", f"{brutto:,.2f} EUR"],
        ]
        summen_table = Table(summen_data, colWidths=[130 * mm, 38 * mm])
        summen_table.setStyle(TableStyle([
            ("ALIGN", (1, 0), (1, -1), "RIGHT"),
            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 11),
            ("LINEABOVE", (0, -1), (-1, -1), 1.5, colors.HexColor("#2d2d2d")),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(summen_table)

        # Hinweise
        story.append(Spacer(1, 10 * mm))
        story.append(Paragraph(
            "Dieses Angebot ist 30 Tage gueltig. "
            "Alle Preise verstehen sich ab Werk Ober-Moerlen. "
            "Montage und Lieferung nach Vereinbarung.",
            styles["BrandSub"],
        ))

        doc.build(story)

        self.logger.info("Angebots-PDF erstellt: %s", pfad)
        return {
            "status": "ok",
            "datei": str(pfad),
            "dateiname": dateiname,
            "angebotspreis_netto": round(netto, 2),
            "angebotspreis_brutto": round(brutto, 2),
        }

    # ------------------------------------------------------------------
    # Interne Kalkulations-PDF
    # ------------------------------------------------------------------

    async def _export_intern_pdf(
        self, kalkulation: dict, projekt_id: str
    ) -> dict[str, Any]:
        """Generiert interne Kalkulations-PDF mit allen Kostendetails."""
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
        )
        from reportlab.lib.styles import getSampleStyleSheet

        dateiname = f"Kalkulation_intern_{projekt_id}_{datetime.now().strftime('%Y%m%d')}.pdf"
        pfad = self._export_verzeichnis / dateiname

        doc = SimpleDocTemplate(
            str(pfad), pagesize=A4,
            leftMargin=15 * mm, rightMargin=15 * mm,
            topMargin=20 * mm, bottomMargin=15 * mm,
        )

        styles = getSampleStyleSheet()
        story: list = []

        zuschlaege = kalkulation.get("zuschlaege", {})
        material = kalkulation.get("materialkosten", {})
        maschinen = kalkulation.get("maschinenkosten", {})
        lohn = kalkulation.get("lohnkosten", {})

        # Header
        story.append(Paragraph(
            f"<b>INTERNE KALKULATION – {projekt_id}</b>", styles["Heading1"]
        ))
        story.append(Paragraph(
            f"Erstellt: {datetime.now().strftime('%d.%m.%Y %H:%M')} | "
            f"Typ: {kalkulation.get('projekt_typ', 'standard')}",
            styles["Normal"],
        ))
        story.append(Spacer(1, 6 * mm))

        # Kostenuebersicht
        story.append(Paragraph("<b>Kostenuebersicht</b>", styles["Heading2"]))

        kosten_data = [
            ["Kostenart", "Betrag (EUR)"],
            ["Materialkosten", f"{material.get('materialkosten_gesamt', 0):,.2f}"],
            ["Maschinenkosten", f"{maschinen.get('maschinenkosten_gesamt', 0):,.2f}"],
            ["Lohnkosten", f"{lohn.get('lohnkosten_gesamt', 0):,.2f}"],
            ["= Herstellkosten", f"{zuschlaege.get('herstellkosten', 0):,.2f}"],
            [f"+ GKZ ({zuschlaege.get('gemeinkosten', {}).get('satz', 0):.0%})",
             f"{zuschlaege.get('gemeinkosten', {}).get('betrag', 0):,.2f}"],
            ["= Selbstkosten", f"{zuschlaege.get('selbstkosten', 0):,.2f}"],
            [f"+ Gewinn ({zuschlaege.get('gewinn', {}).get('satz', 0):.0%})",
             f"{zuschlaege.get('gewinn', {}).get('betrag', 0):,.2f}"],
            [f"+ Wagnis ({zuschlaege.get('wagnis', {}).get('satz', 0):.0%})",
             f"{zuschlaege.get('wagnis', {}).get('betrag', 0):,.2f}"],
            ["+ Montage-Zuschlag",
             f"{zuschlaege.get('montage_zuschlag', {}).get('betrag', 0):,.2f}"],
            ["+ Fremdleistungen",
             f"{zuschlaege.get('fremdleistungen', {}).get('kosten', 0):,.2f}"],
            ["+ FL-Zuschlag",
             f"{zuschlaege.get('fremdleistungen', {}).get('zuschlag', 0):,.2f}"],
            ["= ANGEBOTSPREIS (netto)", f"{zuschlaege.get('angebotspreis_gesamt', 0):,.2f}"],
        ]

        table = Table(kosten_data, colWidths=[110 * mm, 60 * mm])
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#333333")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ALIGN", (1, 0), (1, -1), "RIGHT"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("FONTNAME", (0, 4), (-1, 4), "Helvetica-Bold"),  # Herstellkosten
            ("FONTNAME", (0, 6), (-1, 6), "Helvetica-Bold"),  # Selbstkosten
            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),  # Angebotspreis
            ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#e6f3e6")),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(table)
        story.append(Spacer(1, 6 * mm))

        # Marge
        marge = zuschlaege.get("marge", {})
        story.append(Paragraph(
            f"<b>Marge: {marge.get('absolut', 0):,.2f} EUR "
            f"({marge.get('prozent', 0):.1f}%)</b>",
            styles["Heading3"],
        ))
        story.append(Spacer(1, 6 * mm))

        # Gewerke-Detail (Lohn)
        gewerke = lohn.get("gewerke", {})
        if gewerke:
            story.append(Paragraph("<b>Lohnkosten nach Gewerk</b>", styles["Heading2"]))
            gewerk_data = [["Gewerk", "Stunden", "Kosten (EUR)"]]
            for name, data in gewerke.items():
                gewerk_data.append([
                    name.replace("_", " ").title(),
                    f"{data.get('stunden', 0):.1f}h",
                    f"{data.get('kosten', 0):,.2f}",
                ])
            gt = Table(gewerk_data, colWidths=[80 * mm, 40 * mm, 50 * mm])
            gt.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#555555")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]))
            story.append(gt)
            story.append(Spacer(1, 6 * mm))

        # Maschinen-Zeitplan
        zeitplan = maschinen.get("zeitplan", {})
        if zeitplan:
            story.append(Paragraph("<b>Maschineneinsatz</b>", styles["Heading2"]))
            zeit_data = [["Maschine", "Stunden"]]
            labels = {
                "cnc_stunden": "CNC (Holzher Nextec 7707)",
                "kanten_stunden": "Kantenanleimmaschine",
                "saege_stunden": "Formatkreissaege",
                "bohr_stunden": "Bohrautomat",
            }
            for key, label in labels.items():
                val = zeitplan.get(key, 0)
                if val > 0:
                    zeit_data.append([label, f"{val:.1f}h"])
            cnc_schichten = maschinen.get("cnc_schichten", 0)
            if cnc_schichten > 0:
                zeit_data.append(["CNC-Schichten (a 8h)", str(cnc_schichten)])

            zt = Table(zeit_data, colWidths=[120 * mm, 50 * mm])
            zt.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#555555")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ALIGN", (1, 0), (1, -1), "RIGHT"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]))
            story.append(zt)
            story.append(Spacer(1, 6 * mm))

        # Warnungen
        warnungen = kalkulation.get("warnungen", [])
        if warnungen:
            story.append(Paragraph("<b>Warnungen</b>", styles["Heading2"]))
            for w in warnungen:
                story.append(Paragraph(f"! {w}", styles["Normal"]))
            story.append(Spacer(1, 4 * mm))

        # Footer
        story.append(Paragraph(
            "<i>VERTRAULICH – Nur fuer internen Gebrauch</i>",
            styles["Normal"],
        ))

        doc.build(story)

        self.logger.info("Interne Kalkulations-PDF erstellt: %s", pfad)
        return {"status": "ok", "datei": str(pfad), "dateiname": dateiname}

    # ------------------------------------------------------------------
    # Excel-Export
    # ------------------------------------------------------------------

    async def _export_excel(
        self, kalkulation: dict, projekt_id: str
    ) -> dict[str, Any]:
        """Exportiert Kalkulation als Excel-Datei."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, numbers

        dateiname = f"Kalkulation_{projekt_id}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        pfad = self._export_verzeichnis / dateiname

        wb = Workbook()

        # --- Sheet 1: Uebersicht ---
        ws = wb.active
        ws.title = "Uebersicht"

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill("solid", fgColor="333333")
        bold = Font(bold=True)

        zuschlaege = kalkulation.get("zuschlaege", {})

        uebersicht = [
            ("Kostenart", "Betrag (EUR)"),
            ("Materialkosten", zuschlaege.get("materialkosten", 0)),
            ("Maschinenkosten", zuschlaege.get("maschinenkosten", 0)),
            ("Lohnkosten", zuschlaege.get("lohnkosten", 0)),
            ("Herstellkosten", zuschlaege.get("herstellkosten", 0)),
            ("Gemeinkosten (GKZ)", zuschlaege.get("gemeinkosten", {}).get("betrag", 0)),
            ("Selbstkosten", zuschlaege.get("selbstkosten", 0)),
            ("Gewinn", zuschlaege.get("gewinn", {}).get("betrag", 0)),
            ("Wagnis (VOB)", zuschlaege.get("wagnis", {}).get("betrag", 0)),
            ("Montage-Zuschlag", zuschlaege.get("montage_zuschlag", {}).get("betrag", 0)),
            ("Fremdleistungen", zuschlaege.get("fremdleistungen", {}).get("kosten", 0)),
            ("FL-Zuschlag", zuschlaege.get("fremdleistungen", {}).get("zuschlag", 0)),
            ("ANGEBOTSPREIS (netto)", zuschlaege.get("angebotspreis_gesamt", 0)),
        ]

        for row_idx, (label, value) in enumerate(uebersicht, 1):
            ws.cell(row=row_idx, column=1, value=label)
            cell = ws.cell(row=row_idx, column=2, value=value)
            if row_idx == 1:
                ws.cell(row=1, column=1).font = header_font
                ws.cell(row=1, column=1).fill = header_fill
                ws.cell(row=1, column=2).font = header_font
                ws.cell(row=1, column=2).fill = header_fill
            elif isinstance(value, (int, float)):
                cell.number_format = '#,##0.00'
            if label in ("Herstellkosten", "Selbstkosten", "ANGEBOTSPREIS (netto)"):
                ws.cell(row=row_idx, column=1).font = bold
                ws.cell(row=row_idx, column=2).font = bold

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 18

        # --- Sheet 2: Positionen ---
        ws2 = wb.create_sheet("Positionen")
        pos_headers = ["Pos.", "Kurztext", "Menge", "Einheit", "Material", "Lackierung", "Fremdleistung"]
        for col_idx, h in enumerate(pos_headers, 1):
            cell = ws2.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill

        positionen = kalkulation.get("positionen", [])
        for row_idx, pos in enumerate(positionen, 2):
            ws2.cell(row=row_idx, column=1, value=pos.get("pos_nr", ""))
            ws2.cell(row=row_idx, column=2, value=pos.get("kurztext", ""))
            ws2.cell(row=row_idx, column=3, value=float(pos.get("menge", 0)))
            ws2.cell(row=row_idx, column=4, value=pos.get("einheit", ""))
            ws2.cell(row=row_idx, column=5, value=pos.get("material", ""))
            ws2.cell(row=row_idx, column=6, value="Ja" if pos.get("ist_lackierung") else "")
            ws2.cell(row=row_idx, column=7, value="Ja" if pos.get("ist_fremdleistung") else "")

        for col in ["A", "B", "C", "D", "E", "F", "G"]:
            ws2.column_dimensions[col].width = 18
        ws2.column_dimensions["B"].width = 40

        # --- Sheet 3: Gewerke ---
        ws3 = wb.create_sheet("Gewerke")
        gewerke = kalkulation.get("lohnkosten", {}).get("gewerke", {})
        gewerk_headers = ["Gewerk", "Stunden", "Kosten (EUR)"]
        for col_idx, h in enumerate(gewerk_headers, 1):
            cell = ws3.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill

        for row_idx, (name, data) in enumerate(gewerke.items(), 2):
            ws3.cell(row=row_idx, column=1, value=name.replace("_", " ").title())
            ws3.cell(row=row_idx, column=2, value=data.get("stunden", 0))
            cell = ws3.cell(row=row_idx, column=3, value=data.get("kosten", 0))
            cell.number_format = '#,##0.00'

        ws3.column_dimensions["A"].width = 25
        ws3.column_dimensions["B"].width = 15
        ws3.column_dimensions["C"].width = 18

        wb.save(str(pfad))

        self.logger.info("Excel-Export erstellt: %s", pfad)
        return {"status": "ok", "datei": str(pfad), "dateiname": dateiname}

    # ------------------------------------------------------------------
    # GAEB-Export (X83 ausgefuellt)
    # ------------------------------------------------------------------

    async def _export_gaeb(
        self, kalkulation: dict, projekt_id: str
    ) -> dict[str, Any]:
        """Exportiert ausgefuellte GAEB-X83 Datei mit Einheitspreisen."""
        from lxml import etree

        dateiname = f"Angebot_{projekt_id}.x83"
        pfad = self._export_verzeichnis / dateiname

        zuschlaege = kalkulation.get("zuschlaege", {})
        positionen = kalkulation.get("positionen", [])
        angebotspreis = zuschlaege.get("angebotspreis_gesamt", 0)

        # GAEB XML aufbauen
        ns = "http://www.gaeb.de/GAEB_DA_XML/200407"
        nsmap = {None: ns}

        gaeb = etree.Element(f"{{{ns}}}GAEB", nsmap=nsmap)
        gaeb.set("Version", "3.1")

        # GAEBInfo
        info = etree.SubElement(gaeb, f"{{{ns}}}GAEBInfo")
        version = etree.SubElement(info, f"{{{ns}}}Version")
        version.text = "3.1"
        prog_info = etree.SubElement(info, f"{{{ns}}}ProgInfo")
        prog_info.text = "Brandstifter Kalkulationstool"

        # Award
        award = etree.SubElement(gaeb, f"{{{ns}}}Award")
        award.set("Dp", "83")

        # BoQ (Bill of Quantities)
        boq = etree.SubElement(award, f"{{{ns}}}BoQ")
        boq_info = etree.SubElement(boq, f"{{{ns}}}BoQInfo")
        name_el = etree.SubElement(boq_info, f"{{{ns}}}Name")
        name_el.text = f"Angebot {projekt_id}"

        boq_body = etree.SubElement(boq, f"{{{ns}}}BoQBody")
        itemlist = etree.SubElement(boq_body, f"{{{ns}}}Itemlist")

        # Positionen schreiben
        preis_pro_pos = angebotspreis / max(len(positionen), 1)

        for pos in positionen:
            menge = float(pos.get("menge", 1))
            ep = preis_pro_pos / max(menge, 1)

            item = etree.SubElement(itemlist, f"{{{ns}}}Item")
            item.set("ID", str(pos.get("pos_nr", "")))

            qty = etree.SubElement(item, f"{{{ns}}}Qty")
            qty.text = f"{menge:.3f}"

            qu = etree.SubElement(item, f"{{{ns}}}QU")
            qu.text = pos.get("einheit", "STK")

            # Beschreibung
            desc = etree.SubElement(item, f"{{{ns}}}Description")
            ot = etree.SubElement(desc, f"{{{ns}}}OutlineText")
            otl = etree.SubElement(ot, f"{{{ns}}}OutlTxt")
            text_el = etree.SubElement(otl, f"{{{ns}}}TextOutlTxt")
            span = etree.SubElement(text_el, f"{{{ns}}}span")
            span.text = pos.get("kurztext", "")

            # Preis
            up = etree.SubElement(item, f"{{{ns}}}UP")
            up.text = f"{ep:.2f}"

            it = etree.SubElement(item, f"{{{ns}}}IT")
            it.text = f"{ep * menge:.2f}"

        tree = etree.ElementTree(gaeb)
        tree.write(str(pfad), xml_declaration=True, encoding="UTF-8", pretty_print=True)

        self.logger.info("GAEB-Export erstellt: %s", pfad)
        return {"status": "ok", "datei": str(pfad), "dateiname": dateiname}

    # ------------------------------------------------------------------
    # Schreiners Buero API
    # ------------------------------------------------------------------

    async def _export_schreiners_buero(
        self, kalkulation: dict, projekt_id: str
    ) -> dict[str, Any]:
        """Sendet Stuecklisten und Preise an Schreiners Buero API."""
        if not self._sb_api_url:
            return {"status": "error", "message": "Schreiners Buero API nicht konfiguriert"}

        import httpx

        positionen = kalkulation.get("positionen", [])
        zuschlaege = kalkulation.get("zuschlaege", {})

        # Payload fuer Schreiners Buero
        sb_data = {
            "projekt": projekt_id,
            "positionen": [
                {
                    "pos_nr": p.get("pos_nr", ""),
                    "bezeichnung": p.get("kurztext", ""),
                    "menge": float(p.get("menge", 0)),
                    "einheit": p.get("einheit", "STK"),
                    "material": p.get("material", ""),
                    "preis": zuschlaege.get("angebotspreis_gesamt", 0) / max(len(positionen), 1),
                }
                for p in positionen
            ],
        }

        try:
            async with httpx.AsyncClient(timeout=30.0) as client:
                r = await client.post(
                    self._sb_api_url,
                    json=sb_data,
                    auth=(self._sb_api_user, self._sb_api_pw) if self._sb_api_user else None,
                )
                return {
                    "status": "ok" if r.status_code < 400 else "error",
                    "status_code": r.status_code,
                    "response": r.text[:500],
                }
        except Exception as exc:
            return {"status": "error", "message": str(exc)}
