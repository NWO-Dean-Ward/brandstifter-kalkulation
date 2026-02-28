"""
AnalyseAgent – Subagent 10.

Analysiert Altprojekte und extrahiert:
- Strukturpreise und Einheitspreise aus Excel/GAEB/PDF
- Stuecklisten und Bauteilmasse aus Smartwop-CSVs
- Maschinenzeiten und Materialpreise
- Inflationsanpassung basierend auf Projektdatum
"""

from __future__ import annotations

import csv
import re
from datetime import datetime
from io import StringIO
from pathlib import Path
from typing import Any

from agents.base_agent import AgentMessage, BaseAgent


# Inflationsrate pro Jahr (konservativ 3-5%)
INFLATIONSRATE_PRO_JAHR = 0.04  # 4% Standard

# Egger-Dekornummern Mapping (haeufig verwendet in Schreinereien)
# Format in CSVs: "W1000_19" -> Dekor W1000, Staerke 19mm
# Format in CSVs: "U750 ST9" -> Dekor U750, Struktur ST9
EGGER_DEKORE: dict[str, dict[str, str]] = {
    # --- Uni-Dekore (U-Serie) ---
    "U104": {"name": "Alabasterweiß", "kategorie": "uni", "typ": "Spanplatte"},
    "U108": {"name": "Cremeweiss", "kategorie": "uni", "typ": "Spanplatte"},
    "U113": {"name": "Weiss", "kategorie": "uni", "typ": "Spanplatte"},
    "U114": {"name": "Sandbeige", "kategorie": "uni", "typ": "Spanplatte"},
    "U156": {"name": "Kieselgrau", "kategorie": "uni", "typ": "Spanplatte"},
    "U201": {"name": "Cremeweiss", "kategorie": "uni", "typ": "Spanplatte"},
    "U222": {"name": "Crema", "kategorie": "uni", "typ": "Spanplatte"},
    "U311": {"name": "Burgundrot", "kategorie": "uni", "typ": "Spanplatte"},
    "U525": {"name": "Moosgruen", "kategorie": "uni", "typ": "Spanplatte"},
    "U540": {"name": "Salbeigruen", "kategorie": "uni", "typ": "Spanplatte"},
    "U599": {"name": "Indigo", "kategorie": "uni", "typ": "Spanplatte"},
    "U600": {"name": "Schwarz", "kategorie": "uni", "typ": "Spanplatte"},
    "U617": {"name": "Anthrazitgrau", "kategorie": "uni", "typ": "Spanplatte"},
    "U636": {"name": "Lavagrau", "kategorie": "uni", "typ": "Spanplatte"},
    "U702": {"name": "Kaschmirgrau", "kategorie": "uni", "typ": "Spanplatte"},
    "U707": {"name": "Dunkelgrau", "kategorie": "uni", "typ": "Spanplatte"},
    "U708": {"name": "Hellgrau", "kategorie": "uni", "typ": "Spanplatte"},
    "U727": {"name": "Steingrau", "kategorie": "uni", "typ": "Spanplatte"},
    "U732": {"name": "Stahlgrau", "kategorie": "uni", "typ": "Spanplatte"},
    "U748": {"name": "Trüffelbraun", "kategorie": "uni", "typ": "Spanplatte"},
    "U750": {"name": "Taupe", "kategorie": "uni", "typ": "Spanplatte"},
    "U755": {"name": "Dakar", "kategorie": "uni", "typ": "Spanplatte"},
    "U763": {"name": "Perlweiß", "kategorie": "uni", "typ": "Spanplatte"},
    "U775": {"name": "Weissbeton", "kategorie": "uni", "typ": "Spanplatte"},
    "U960": {"name": "Graphitschwarz", "kategorie": "uni", "typ": "Spanplatte"},
    "U961": {"name": "Schwarzbraun", "kategorie": "uni", "typ": "Spanplatte"},
    "U999": {"name": "Schwarz", "kategorie": "uni", "typ": "Spanplatte"},
    # --- Weiss-Serie (W) ---
    "W908": {"name": "Premium Weiss", "kategorie": "uni", "typ": "Spanplatte"},
    "W980": {"name": "Platinweiss", "kategorie": "uni", "typ": "Spanplatte"},
    "W1000": {"name": "Premium Weiss", "kategorie": "uni", "typ": "Spanplatte"},
    "W1001": {"name": "Kieselgrau", "kategorie": "uni", "typ": "Spanplatte"},
    "W1100": {"name": "Alpinweiss", "kategorie": "uni", "typ": "Spanplatte"},
    "W1200": {"name": "Polarweiss", "kategorie": "uni", "typ": "Spanplatte"},
    # --- Holzdekore (H-Serie) ---
    "H1101": {"name": "Whitewood", "kategorie": "holzdekor", "typ": "Spanplatte"},
    "H1107": {"name": "Pinie Admont braun", "kategorie": "holzdekor", "typ": "Spanplatte"},
    "H1133": {"name": "Eiche Davos Trüffel", "kategorie": "holzdekor", "typ": "Spanplatte"},
    "H1145": {"name": "Eiche Bardolino natur", "kategorie": "holzdekor", "typ": "Spanplatte"},
    "H1150": {"name": "Eiche Arizona", "kategorie": "holzdekor", "typ": "Spanplatte"},
    "H1176": {"name": "Eiche Halifax Weiss", "kategorie": "holzdekor", "typ": "Spanplatte"},
    "H1180": {"name": "Eiche Halifax natur", "kategorie": "holzdekor", "typ": "Spanplatte"},
    "H1181": {"name": "Eiche Halifax Tabak", "kategorie": "holzdekor", "typ": "Spanplatte"},
    "H1212": {"name": "Eiche Wadi", "kategorie": "holzdekor", "typ": "Spanplatte"},
    "H1250": {"name": "Eiche Navarra", "kategorie": "holzdekor", "typ": "Spanplatte"},
    "H1277": {"name": "Eiche Lakeland hell", "kategorie": "holzdekor", "typ": "Spanplatte"},
    "H1313": {"name": "Eiche Whiteriver sand", "kategorie": "holzdekor", "typ": "Spanplatte"},
    "H1318": {"name": "Eiche Whiteriver braun", "kategorie": "holzdekor", "typ": "Spanplatte"},
    "H1334": {"name": "Eiche Ferrara natur", "kategorie": "holzdekor", "typ": "Spanplatte"},
    "H1336": {"name": "Eiche Ferrara schwarz", "kategorie": "holzdekor", "typ": "Spanplatte"},
    "H1344": {"name": "Eiche Sherman cognac", "kategorie": "holzdekor", "typ": "Spanplatte"},
    "H1345": {"name": "Eiche Sherman natur", "kategorie": "holzdekor", "typ": "Spanplatte"},
    "H1346": {"name": "Eiche Sherman hell", "kategorie": "holzdekor", "typ": "Spanplatte"},
    "H1377": {"name": "Eiche Denver", "kategorie": "holzdekor", "typ": "Spanplatte"},
    "H1387": {"name": "Eiche Huston natur", "kategorie": "holzdekor", "typ": "Spanplatte"},
    "H1399": {"name": "Eiche Thermoform", "kategorie": "holzdekor", "typ": "Spanplatte"},
    "H1401": {"name": "Eiche Casella Bernstein", "kategorie": "holzdekor", "typ": "Spanplatte"},
    "H1486": {"name": "Eiche Pasadena natur", "kategorie": "holzdekor", "typ": "Spanplatte"},
    "H1487": {"name": "Eiche Bern braun", "kategorie": "holzdekor", "typ": "Spanplatte"},
    "H1511": {"name": "Eiche Bavaria Bernstein", "kategorie": "holzdekor", "typ": "Spanplatte"},
    "H1521": {"name": "Eiche Almington natur", "kategorie": "holzdekor", "typ": "Spanplatte"},
    "H1530": {"name": "Eiche Lincoln", "kategorie": "holzdekor", "typ": "Spanplatte"},
    "H1615": {"name": "Eiche Peyronade", "kategorie": "holzdekor", "typ": "Spanplatte"},
    "H1636": {"name": "Eiche Knotty natur", "kategorie": "holzdekor", "typ": "Spanplatte"},
    "H1700": {"name": "Eiche Yukon", "kategorie": "holzdekor", "typ": "Spanplatte"},
    "H1710": {"name": "Treibholz", "kategorie": "holzdekor", "typ": "Spanplatte"},
    "H1733": {"name": "Eiche Mainau", "kategorie": "holzdekor", "typ": "Spanplatte"},
    "H3058": {"name": "Akazie Laredo", "kategorie": "holzdekor", "typ": "Spanplatte"},
    "H3090": {"name": "Fleetwood weiss", "kategorie": "holzdekor", "typ": "Spanplatte"},
    "H3131": {"name": "Walnuss Amore", "kategorie": "holzdekor", "typ": "Spanplatte"},
    "H3133": {"name": "Eiche Davos Trueffel", "kategorie": "holzdekor", "typ": "Spanplatte"},
    "H3146": {"name": "Nussbaum Dijon", "kategorie": "holzdekor", "typ": "Spanplatte"},
    "H3154": {"name": "Eiche Charleston dunkel", "kategorie": "holzdekor", "typ": "Spanplatte"},
    "H3170": {"name": "Eiche Kendal cognac", "kategorie": "holzdekor", "typ": "Spanplatte"},
    "H3176": {"name": "Eiche Halifax Weiss", "kategorie": "holzdekor", "typ": "Spanplatte"},
    "H3303": {"name": "Eiche Hamilton natur", "kategorie": "holzdekor", "typ": "Spanplatte"},
    "H3309": {"name": "Eiche Gladstone Sand", "kategorie": "holzdekor", "typ": "Spanplatte"},
    "H3325": {"name": "Eiche Gladstone Tabak", "kategorie": "holzdekor", "typ": "Spanplatte"},
    "H3326": {"name": "Eiche Gladstone grau", "kategorie": "holzdekor", "typ": "Spanplatte"},
    "H3331": {"name": "Eiche Nebraska natur", "kategorie": "holzdekor", "typ": "Spanplatte"},
    "H3332": {"name": "Eiche Nebraska grau", "kategorie": "holzdekor", "typ": "Spanplatte"},
    "H3342": {"name": "Eiche Olchon Sand", "kategorie": "holzdekor", "typ": "Spanplatte"},
    "H3344": {"name": "Eiche Corbridge natur", "kategorie": "holzdekor", "typ": "Spanplatte"},
    "H3395": {"name": "Eiche Henley", "kategorie": "holzdekor", "typ": "Spanplatte"},
    "H3398": {"name": "Eiche Bergamo", "kategorie": "holzdekor", "typ": "Spanplatte"},
    "H3700": {"name": "Eiche Natur Bern", "kategorie": "holzdekor", "typ": "Spanplatte"},
    "H3710": {"name": "Eiche Natur Yukon", "kategorie": "holzdekor", "typ": "Spanplatte"},
    "H3730": {"name": "Eiche Hickory natur", "kategorie": "holzdekor", "typ": "Spanplatte"},
    "H3760": {"name": "Eiche Lancelot", "kategorie": "holzdekor", "typ": "Spanplatte"},
    # --- Stein/Beton-Dekore ---
    "F186": {"name": "Beton Chicago hellgrau", "kategorie": "steindekor", "typ": "Spanplatte"},
    "F187": {"name": "Beton Chicago dunkelgrau", "kategorie": "steindekor", "typ": "Spanplatte"},
    "F206": {"name": "Pietra Grigia anthrazit", "kategorie": "steindekor", "typ": "Spanplatte"},
    "F274": {"name": "Beton hell", "kategorie": "steindekor", "typ": "Spanplatte"},
    "F275": {"name": "Beton mittel", "kategorie": "steindekor", "typ": "Spanplatte"},
    "F283": {"name": "Titanit anthrazit", "kategorie": "steindekor", "typ": "Spanplatte"},
    "F385": {"name": "Granit Galizia grau", "kategorie": "steindekor", "typ": "Spanplatte"},
    "F462": {"name": "Marmor Paladina hell", "kategorie": "steindekor", "typ": "Spanplatte"},
    "F501": {"name": "Beton geschliffen", "kategorie": "steindekor", "typ": "Spanplatte"},
    "F509": {"name": "Limestone", "kategorie": "steindekor", "typ": "Spanplatte"},
    "F637": {"name": "Chromix Anthrazit", "kategorie": "steindekor", "typ": "Spanplatte"},
    "F638": {"name": "Chromix Silber", "kategorie": "steindekor", "typ": "Spanplatte"},
    "F642": {"name": "Chromix Bronze", "kategorie": "steindekor", "typ": "Spanplatte"},
    "F812": {"name": "Levanto Marmor schwarz", "kategorie": "steindekor", "typ": "Spanplatte"},
    # --- MDF-Spezial ---
    "L598": {"name": "Arktisweiss", "kategorie": "uni", "typ": "MDF"},
}

# Egger-Oberflaechen (Strukturen)
EGGER_STRUKTUREN: dict[str, str] = {
    "ST2": "Perlstruktur matt",
    "ST9": "Feelwood naturpore",
    "ST10": "Deepskin robust",
    "ST12": "Struktur Premium matt",
    "ST15": "Smooth matt",
    "ST19": "Deepskin quer",
    "ST22": "Deepskin längs",
    "ST28": "Diamant",
    "ST33": "Pore matt",
    "ST37": "Feelwood fein",
    "ST38": "Feelwood Altholz",
    "ST76": "Samtstruktur",
    "ST87": "Riffholz",
    "ST89": "Texwood",
    "SM": "Glatt matt",
    "PE": "Perlstruktur",
}


class AnalyseAgent(BaseAgent):
    """Analysiert Altprojekte und extrahiert Preisstrukturen."""

    def __init__(self) -> None:
        super().__init__(name="analyse_agent")
        self._altprojekt_pfad = ""
        self._fragen: list[dict[str, str]] = []

    def configure(self, altprojekt_basis_pfad: str = "") -> None:
        """Konfiguriert den Basispfad fuer Altprojekte."""
        self._altprojekt_pfad = altprojekt_basis_pfad

    async def process(self, message: AgentMessage) -> AgentMessage:
        msg_type = message.msg_type
        handler_map = {
            "scan_altprojekt": self._scan_altprojekt,
            "analyse_excel_lv": self._analyse_excel_lv,
            "analyse_gaeb": self._analyse_gaeb,
            "analyse_smartwop_csvs": self._analyse_smartwop_csvs,
            "analyse_komplett": self._analyse_komplett,
            "inflationsanpassung": self._inflationsanpassung,
            "fragen_abrufen": self._fragen_abrufen,
            "fragen_beantworten": self._fragen_beantworten,
        }

        handler = handler_map.get(msg_type)
        if handler is None:
            return message.create_error(
                sender=self.name,
                error_msg=f"Unbekannter Analyse-Typ: {msg_type}",
            )

        result = await handler(message.payload)
        return message.create_response(sender=self.name, payload=result)

    # ------------------------------------------------------------------
    # 1. Ordner scannen
    # ------------------------------------------------------------------

    async def _scan_altprojekt(self, payload: dict) -> dict[str, Any]:
        """Scannt einen Altprojekt-Ordner und listet alle Dateien."""
        pfad = Path(payload.get("pfad", self._altprojekt_pfad))
        if not pfad.exists():
            return {"status": "error", "message": f"Pfad nicht gefunden: {pfad}"}

        dateien: list[dict[str, Any]] = []
        for f in sorted(pfad.rglob("*")):
            if f.is_file() and not f.name.startswith("."):
                suffix = f.suffix.lower()
                dateien.append({
                    "pfad": str(f),
                    "name": f.name,
                    "ordner": str(f.parent.relative_to(pfad)),
                    "typ": suffix,
                    "groesse_kb": round(f.stat().st_size / 1024, 1),
                })

        # Nach Typ gruppieren
        typen: dict[str, int] = {}
        for d in dateien:
            typen[d["typ"]] = typen.get(d["typ"], 0) + 1

        return {
            "status": "ok",
            "pfad": str(pfad),
            "dateien_gesamt": len(dateien),
            "typen": typen,
            "dateien": dateien,
            "analysierbare": [d for d in dateien if d["typ"] in (
                ".csv", ".xlsx", ".xls", ".d83", ".d84", ".x83", ".x84", ".pdf"
            )],
        }

    # ------------------------------------------------------------------
    # 2. Excel-LV analysieren
    # ------------------------------------------------------------------

    async def _analyse_excel_lv(self, payload: dict) -> dict[str, Any]:
        """Extrahiert Positionen und Preise aus einer Excel-Kalkulation."""
        pfad = payload.get("pfad", "")
        if not pfad or not Path(pfad).exists():
            return {"status": "error", "message": f"Datei nicht gefunden: {pfad}"}

        from openpyxl import load_workbook

        wb = load_workbook(pfad, read_only=True, data_only=True)
        positionen: list[dict[str, Any]] = []
        bereiche: list[dict[str, Any]] = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            # Header erkennen
            header = [str(c).strip().lower() if c else "" for c in rows[0]]

            # OZ/Pos-Nr und EP/GP Spalten finden
            oz_idx = self._find_col(header, ["oz", "pos", "pos-nr", "pos_nr", "position"])
            ep_idx = self._find_col(header, ["ep", "einheitspreis", "e.p."])
            gp_idx = self._find_col(header, ["gp", "gesamtpreis", "g.p.", "gesamt"])
            text_idx = self._find_col(header, ["kurztext", "bezeichnung", "text", "beschreibung"])
            menge_idx = self._find_col(header, ["menge", "anzahl", "qty"])
            einheit_idx = self._find_col(header, ["einheit", "me", "einh"])
            art_idx = self._find_col(header, ["posart", "art", "typ"])

            for row_data in rows[1:]:
                if not row_data or all(c is None for c in row_data):
                    continue

                oz = str(row_data[oz_idx] or "").strip() if oz_idx is not None else ""
                art = str(row_data[art_idx] or "").strip() if art_idx is not None else ""

                if art.upper() == "BEREICH":
                    text = str(row_data[text_idx] or "").strip() if text_idx is not None else ""
                    bereiche.append({"oz": oz, "bezeichnung": text})
                    continue

                if art.upper() in ("HINWEIS", ""):
                    if not oz or oz.upper() in ("HINWEIS", ""):
                        continue

                ep = self._safe_float(row_data[ep_idx]) if ep_idx is not None else 0
                gp = self._safe_float(row_data[gp_idx]) if gp_idx is not None else 0
                menge = self._safe_float(row_data[menge_idx]) if menge_idx is not None else 0
                text = str(row_data[text_idx] or "").strip() if text_idx is not None else ""
                einheit = str(row_data[einheit_idx] or "STK").strip() if einheit_idx is not None else "STK"

                if ep > 0 or gp > 0:
                    positionen.append({
                        "pos_nr": oz,
                        "kurztext": text,
                        "menge": menge,
                        "einheit": einheit,
                        "einheitspreis": ep,
                        "gesamtpreis": gp if gp > 0 else ep * menge,
                    })

        wb.close()

        return {
            "status": "ok",
            "datei": pfad,
            "sheet_namen": wb.sheetnames if hasattr(wb, 'sheetnames') else [],
            "bereiche": bereiche,
            "positionen": positionen,
            "positionen_anzahl": len(positionen),
            "gesamtsumme": sum(p["gesamtpreis"] for p in positionen),
        }

    # ------------------------------------------------------------------
    # 3. GAEB analysieren
    # ------------------------------------------------------------------

    async def _analyse_gaeb(self, payload: dict) -> dict[str, Any]:
        """Extrahiert Positionen aus einer GAEB-Datei (D83/D84/X83/X84)."""
        pfad = payload.get("pfad", "")
        if not pfad or not Path(pfad).exists():
            return {"status": "error", "message": f"Datei nicht gefunden: {pfad}"}

        from lxml import etree

        tree = etree.parse(pfad)
        root = tree.getroot()
        ns = {"g": "http://www.gaeb.de/GAEB_DA_XML/200407"}

        positionen: list[dict[str, Any]] = []

        for item in root.findall(".//g:Item", ns):
            pos_nr = item.get("ID", "")
            qty_el = item.find("g:Qty", ns)
            qu_el = item.find("g:QU", ns)
            up_el = item.find("g:UP", ns)
            it_el = item.find("g:IT", ns)

            menge = float(qty_el.text) if qty_el is not None and qty_el.text else 0
            einheit = qu_el.text if qu_el is not None and qu_el.text else "STK"
            ep = float(up_el.text) if up_el is not None and up_el.text else 0
            gp = float(it_el.text) if it_el is not None and it_el.text else 0

            # Kurztext extrahieren
            kurztext = ""
            for span in item.findall(".//g:span", ns):
                if span.text:
                    kurztext += span.text + " "
            kurztext = kurztext.strip()

            positionen.append({
                "pos_nr": pos_nr,
                "kurztext": kurztext,
                "menge": menge,
                "einheit": einheit,
                "einheitspreis": ep,
                "gesamtpreis": gp if gp > 0 else ep * menge,
            })

        return {
            "status": "ok",
            "datei": pfad,
            "format": Path(pfad).suffix.upper(),
            "positionen": positionen,
            "positionen_anzahl": len(positionen),
            "gesamtsumme": sum(p["gesamtpreis"] for p in positionen),
            "hat_preise": any(p["einheitspreis"] > 0 for p in positionen),
        }

    # ------------------------------------------------------------------
    # 4. Smartwop-CSV Stuecklisten analysieren
    # ------------------------------------------------------------------

    async def _analyse_smartwop_csvs(self, payload: dict) -> dict[str, Any]:
        """Analysiert Smartwop-Export-CSVs (Stuecklisten pro Moebel)."""
        pfad = Path(payload.get("pfad", ""))
        if not pfad.exists():
            return {"status": "error", "message": f"Pfad nicht gefunden: {pfad}"}

        csv_dateien = list(pfad.rglob("*.csv"))
        if not csv_dateien:
            return {"status": "error", "message": "Keine CSV-Dateien gefunden"}

        moebel_stuecklisten: list[dict[str, Any]] = []
        alle_materialien: dict[str, int] = {}
        alle_beschlaege: list[dict[str, Any]] = []

        for csv_datei in sorted(csv_dateien):
            stueckliste = self._parse_smartwop_csv(csv_datei)
            if stueckliste:
                moebel_stuecklisten.append(stueckliste)

                for bauteil in stueckliste.get("bauteile", []):
                    mat = bauteil.get("material_code", "")
                    if mat:
                        alle_materialien[mat] = alle_materialien.get(mat, 0) + bauteil.get("anzahl", 1)

                alle_beschlaege.extend(stueckliste.get("beschlaege", []))

        # Materialien mit Egger-Namen anreichern
        materialien_details: list[dict[str, Any]] = []
        for code, anzahl in sorted(alle_materialien.items()):
            egger = self._resolve_egger_code(code)
            materialien_details.append({
                "code": code,
                "anzahl": anzahl,
                "egger_name": egger.get("egger_name", ""),
                "egger_kategorie": egger.get("egger_kategorie", ""),
                "staerke_mm": egger.get("staerke_mm", 0),
                "unbekannt": egger.get("egger_unbekannt", False),
            })

        return {
            "status": "ok",
            "pfad": str(pfad),
            "csv_dateien": len(csv_dateien),
            "moebel_stuecklisten": moebel_stuecklisten,
            "materialien_uebersicht": alle_materialien,
            "materialien_details": materialien_details,
            "beschlaege_gesamt": len(alle_beschlaege),
            "fragen": self._fragen,
        }

    def _parse_smartwop_csv(self, csv_datei: Path) -> dict[str, Any] | None:
        """Parst eine einzelne Smartwop-CSV Stueckliste."""
        try:
            # Smartwop CSVs koennen verschiedene Encodings haben
            for encoding in ("utf-8-sig", "iso-8859-1", "cp1252"):
                try:
                    content = csv_datei.read_text(encoding=encoding)
                    break
                except UnicodeDecodeError:
                    continue
            else:
                return None

            reader = csv.reader(StringIO(content), delimiter=";")
            rows = list(reader)
            if not rows:
                return None

            # Erste Zeile = Moebel-Header (leer;Name;Anzahl;Laenge;Breite;Tiefe;...)
            header_row = rows[0]
            moebel_name = header_row[1] if len(header_row) > 1 else csv_datei.stem
            moebel_laenge = self._parse_de_zahl(header_row[3]) if len(header_row) > 3 else 0
            moebel_breite = self._parse_de_zahl(header_row[4]) if len(header_row) > 4 else 0
            moebel_tiefe = self._parse_de_zahl(header_row[5]) if len(header_row) > 5 else 0

            bauteile: list[dict[str, Any]] = []
            beschlaege: list[dict[str, Any]] = []

            for row in rows[1:]:
                if len(row) < 3:
                    continue

                material_code = row[0].strip()
                bezeichnung = row[1].strip() if len(row) > 1 else ""
                anzahl = int(self._parse_de_zahl(row[2])) if len(row) > 2 and row[2].strip() else 1
                laenge = self._parse_de_zahl(row[3]) if len(row) > 3 else 0
                breite = self._parse_de_zahl(row[4]) if len(row) > 4 else 0

                # Unterscheidung: Bauteil (hat Masse) vs Beschlag (keine Masse)
                if laenge > 0 or breite > 0:
                    # Kanten extrahieren (Spalten 5-8)
                    kanten = []
                    for ki in range(5, 9):
                        if len(row) > ki and row[ki].strip():
                            kanten.append(row[ki].strip())

                    # HOP-Datei Referenz finden (typisch Spalte 14+)
                    hop_datei = ""
                    for col in row:
                        if col.strip().endswith(".hop"):
                            hop_datei = col.strip()
                            break

                    # Egger-Dekornummer aufloesen
                    egger_info = self._resolve_egger_code(material_code)

                    bauteile.append({
                        "material_code": material_code,
                        "bezeichnung": bezeichnung,
                        "anzahl": anzahl,
                        "laenge_mm": laenge,
                        "breite_mm": breite,
                        "kanten": kanten,
                        "hop_datei": hop_datei,
                        **egger_info,
                    })
                elif material_code and bezeichnung:
                    beschlaege.append({
                        "artikel_nr": material_code,
                        "bezeichnung": bezeichnung,
                        "anzahl": anzahl,
                    })

            return {
                "datei": str(csv_datei),
                "moebel_name": moebel_name,
                "moebel_masse": {
                    "laenge_mm": moebel_laenge,
                    "breite_mm": moebel_breite,
                    "tiefe_mm": moebel_tiefe,
                },
                "bauteile": bauteile,
                "bauteile_anzahl": len(bauteile),
                "beschlaege": beschlaege,
                "beschlaege_anzahl": len(beschlaege),
                "ordner": csv_datei.parent.name,
            }
        except Exception as exc:
            self.logger.warning("CSV-Parse-Fehler %s: %s", csv_datei, exc)
            return None

    # ------------------------------------------------------------------
    # 5. Komplett-Analyse (alles zusammen)
    # ------------------------------------------------------------------

    async def _analyse_komplett(self, payload: dict) -> dict[str, Any]:
        """Fuehrt eine vollstaendige Analyse eines Altprojekt-Ordners durch."""
        pfad = Path(payload.get("pfad", self._altprojekt_pfad))
        if not pfad.exists():
            return {"status": "error", "message": f"Pfad nicht gefunden: {pfad}"}

        self._fragen.clear()
        ergebnis: dict[str, Any] = {"pfad": str(pfad), "status": "ok"}

        # 1. Scan
        scan = await self._scan_altprojekt({"pfad": str(pfad)})
        ergebnis["scan"] = scan

        # 2. Excel-LVs analysieren
        excel_ergebnisse = []
        for d in scan.get("analysierbare", []):
            if d["typ"] in (".xlsx", ".xls"):
                try:
                    r = await self._analyse_excel_lv({"pfad": d["pfad"]})
                    excel_ergebnisse.append(r)
                except Exception as exc:
                    excel_ergebnisse.append({"status": "error", "datei": d["pfad"], "message": str(exc)})
        ergebnis["excel_analysen"] = excel_ergebnisse

        # 3. GAEB-Dateien analysieren
        gaeb_ergebnisse = []
        for d in scan.get("analysierbare", []):
            if d["typ"] in (".d83", ".d84", ".x83", ".x84"):
                try:
                    r = await self._analyse_gaeb({"pfad": d["pfad"]})
                    gaeb_ergebnisse.append(r)
                except Exception as exc:
                    gaeb_ergebnisse.append({"status": "error", "datei": d["pfad"], "message": str(exc)})
        ergebnis["gaeb_analysen"] = gaeb_ergebnisse

        # 4. Smartwop-CSVs analysieren
        smartwop_dirs = set()
        for d in scan.get("analysierbare", []):
            if d["typ"] == ".csv":
                smartwop_dirs.add(str(Path(d["pfad"]).parent))

        smartwop_ergebnisse = []
        for sw_dir in sorted(smartwop_dirs):
            try:
                r = await self._analyse_smartwop_csvs({"pfad": sw_dir})
                smartwop_ergebnisse.append(r)
            except Exception as exc:
                smartwop_ergebnisse.append({"status": "error", "pfad": sw_dir, "message": str(exc)})
        ergebnis["smartwop_analysen"] = smartwop_ergebnisse

        # 5. Fragen sammeln
        ergebnis["offene_fragen"] = list(self._fragen)

        return ergebnis

    # ------------------------------------------------------------------
    # 6. Inflationsanpassung
    # ------------------------------------------------------------------

    async def _inflationsanpassung(self, payload: dict) -> dict[str, Any]:
        """Berechnet Inflationsanpassung fuer Altpreise."""
        preise = payload.get("preise", [])
        projekt_datum_str = payload.get("projekt_datum", "")
        rate = payload.get("rate_pro_jahr", INFLATIONSRATE_PRO_JAHR)

        if not projekt_datum_str:
            return {"status": "error", "message": "Kein Projektdatum angegeben"}

        try:
            projekt_datum = datetime.fromisoformat(projekt_datum_str)
        except ValueError:
            # Versuche deutsches Format
            try:
                projekt_datum = datetime.strptime(projekt_datum_str, "%d.%m.%Y")
            except ValueError:
                return {"status": "error", "message": f"Datumsformat nicht erkannt: {projekt_datum_str}"}

        heute = datetime.now()
        jahre = (heute - projekt_datum).days / 365.25
        faktor = (1 + rate) ** jahre

        angepasste_preise = []
        for p in preise:
            original = float(p.get("preis", 0))
            angepasst = round(original * faktor, 2)
            angepasste_preise.append({
                **p,
                "original_preis": original,
                "angepasster_preis": angepasst,
                "aufschlag_prozent": round((faktor - 1) * 100, 1),
            })

        return {
            "status": "ok",
            "projekt_datum": projekt_datum_str,
            "jahre_seit_projekt": round(jahre, 1),
            "inflationsrate": rate,
            "faktor": round(faktor, 4),
            "aufschlag_prozent": round((faktor - 1) * 100, 1),
            "preise": angepasste_preise,
        }

    # ------------------------------------------------------------------
    # 7. Fragen-System (bei unklaren Felddefinitionen)
    # ------------------------------------------------------------------

    async def _fragen_abrufen(self, payload: dict) -> dict[str, Any]:
        """Gibt alle offenen Fragen zurueck."""
        return {"fragen": list(self._fragen), "anzahl": len(self._fragen)}

    async def _fragen_beantworten(self, payload: dict) -> dict[str, Any]:
        """Nimmt Antworten auf offene Fragen entgegen."""
        antworten = payload.get("antworten", {})
        beantwortet = 0
        for frage in self._fragen:
            frage_id = frage.get("id", "")
            if frage_id in antworten:
                frage["antwort"] = antworten[frage_id]
                frage["status"] = "beantwortet"
                beantwortet += 1

        return {
            "status": "ok",
            "beantwortet": beantwortet,
            "offen": len([f for f in self._fragen if f.get("status") != "beantwortet"]),
        }

    def _add_frage(self, kontext: str, frage_text: str) -> None:
        """Fuegt eine Frage an den User hinzu."""
        self._fragen.append({
            "id": f"frage_{len(self._fragen) + 1}",
            "kontext": kontext,
            "frage": frage_text,
            "status": "offen",
        })

    # ------------------------------------------------------------------
    # Hilfsfunktionen
    # ------------------------------------------------------------------

    @staticmethod
    def _resolve_egger_code(code: str) -> dict[str, Any]:
        """Loest einen Egger-Material-Code auf.

        Formate:
        - "W1000_19"  -> Dekor W1000, Staerke 19mm
        - "U750 ST9"  -> Dekor U750, Struktur ST9
        - "H1345_19"  -> Dekor H1345, Staerke 19mm
        - "F186 ST9"  -> Dekor F186, Struktur ST9
        """
        if not code:
            return {}

        # Staerke extrahieren: "_19" am Ende = 19mm
        staerke_mm = 0.0
        dekor_teil = code.strip()
        struktur = ""

        # Pattern: CODE_STAERKE (z.B. W1000_19)
        match = re.match(r"^([A-Za-z]\d+)[_\-](\d+)$", dekor_teil)
        if match:
            dekor_teil = match.group(1).upper()
            staerke_mm = float(match.group(2))
        else:
            # Pattern: CODE STRUKTUR (z.B. U750 ST9)
            match = re.match(r"^([A-Za-z]\d+)\s+(ST\d+|SM|PE)$", dekor_teil, re.IGNORECASE)
            if match:
                dekor_teil = match.group(1).upper()
                struktur = match.group(2).upper()
            else:
                # Nur Dekor-Code ohne Zusatz
                match = re.match(r"^([A-Za-z]\d+)$", dekor_teil)
                if match:
                    dekor_teil = match.group(1).upper()
                else:
                    return {"egger_unbekannt": True, "raw_code": code}

        info: dict[str, Any] = {"egger_dekor": dekor_teil}

        if staerke_mm > 0:
            info["staerke_mm"] = staerke_mm

        # Dekor im Katalog nachschlagen
        dekor_data = EGGER_DEKORE.get(dekor_teil)
        if dekor_data:
            info["egger_name"] = dekor_data["name"]
            info["egger_kategorie"] = dekor_data["kategorie"]
            info["egger_typ"] = dekor_data["typ"]
        else:
            info["egger_unbekannt"] = True

        # Struktur aufloesen
        if struktur:
            info["egger_struktur"] = struktur
            info["egger_struktur_name"] = EGGER_STRUKTUREN.get(struktur, "")

        return info

    @staticmethod
    def _find_col(header: list[str], candidates: list[str]) -> int | None:
        """Findet einen Spaltenindex anhand mehrerer moeglicher Header-Namen."""
        for i, h in enumerate(header):
            for c in candidates:
                if c in h:
                    return i
        return None

    @staticmethod
    def _safe_float(val: Any) -> float:
        """Konvertiert einen Wert sicher zu float."""
        if val is None:
            return 0.0
        if isinstance(val, (int, float)):
            return float(val)
        s = str(val).strip().replace(".", "").replace(",", ".")
        try:
            return float(s)
        except ValueError:
            return 0.0

    @staticmethod
    def _parse_de_zahl(val: str) -> float:
        """Parst eine deutsche Zahl (1.234,56 -> 1234.56)."""
        if not val or not val.strip():
            return 0.0
        s = val.strip().replace(".", "").replace(",", ".")
        try:
            return float(s)
        except ValueError:
            return 0.0
