"""
DokumentParser – Subagent 1.

Liest und interpretiert Ausschreibungsdokumente:
- GAEB-Dateien (.x83, .x84 = XML-basiert)
- GAEB-Dateien (.d83 = aelteres Textformat)
- PDF (via pdfplumber)
- Excel (.xlsx)

Extrahiert LV-Positionen in ein einheitliches internes Format.
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from agents.base_agent import AgentMessage, BaseAgent


class DokumentParser(BaseAgent):
    """Parst Ausschreibungsdokumente und extrahiert LV-Positionen."""

    def __init__(self) -> None:
        super().__init__(name="dokument_parser")

    async def process(self, message: AgentMessage) -> AgentMessage:
        """Verarbeitet parse_dokument-Anfragen."""
        datei_pfad = message.payload.get("datei_pfad", "")
        raw_text = message.payload.get("raw_text", "")

        if datei_pfad:
            pfad = Path(datei_pfad)
            if not pfad.exists():
                return message.create_error(
                    sender=self.name,
                    error_msg=f"Datei nicht gefunden: {datei_pfad}",
                )

            datei_typ = self._detect_type(pfad)

            if datei_typ == "gaeb_xml":
                positionen = self._parse_gaeb_xml(pfad)
            elif datei_typ == "gaeb_d83":
                positionen = self._parse_gaeb_d83(pfad)
            elif datei_typ == "pdf":
                positionen = self._parse_pdf(pfad)
            elif datei_typ in ("xlsx", "xls"):
                positionen = self._parse_excel(pfad)
            else:
                return message.create_error(
                    sender=self.name,
                    error_msg=f"Nicht unterstuetzter Dateityp: {pfad.suffix}",
                )
        elif raw_text:
            positionen = self._parse_freitext(raw_text)
        else:
            return message.create_error(
                sender=self.name,
                error_msg="Weder datei_pfad noch raw_text im Payload.",
            )

        # Nachbearbeitung
        positionen = self._erkennung_lackierung(positionen)
        projekt_typ = self._detect_projekt_typ(positionen, message.payload)

        return message.create_response(
            sender=self.name,
            payload={
                "positionen": positionen,
                "projekt_typ": projekt_typ,
                "anzahl_positionen": len(positionen),
                "quelle": datei_pfad or "freitext",
            },
        )

    # ------------------------------------------------------------------
    # Dateityp-Erkennung
    # ------------------------------------------------------------------

    def _detect_type(self, pfad: Path) -> str:
        suffix = pfad.suffix.lower()
        if suffix in (".x83", ".x84"):
            return "gaeb_xml"
        if suffix in (".d83", ".d84", ".p83", ".p84"):
            return "gaeb_d83"
        if suffix == ".pdf":
            return "pdf"
        if suffix in (".xlsx", ".xls"):
            return "xlsx"
        return suffix.lstrip(".")

    # ------------------------------------------------------------------
    # GAEB XML Parser (.x83 / .x84)
    # ------------------------------------------------------------------

    def _parse_gaeb_xml(self, pfad: Path) -> list[dict[str, Any]]:
        """Parst GAEB-XML Dateien (X83/X84 Format).

        GAEB XML Struktur:
        <GAEB>
          <Award>
            <BoQ>          (Bill of Quantities)
              <BoQBody>
                <BoQCtgy>  (Kategorie/Titel)
                  <BoQBody>
                    <Itemlist>
                      <Item> (Position)
        """
        from lxml import etree

        self.logger.info("GAEB-XML parsen: %s", pfad)

        tree = etree.parse(str(pfad))
        root = tree.getroot()

        # Namespace erkennen
        nsmap = root.nsmap
        ns = nsmap.get(None, "")
        if ns:
            ns = f"{{{ns}}}"
        else:
            ns = ""

        positionen: list[dict[str, Any]] = []

        # Alle Item-Elemente finden (rekursiv)
        for item in root.iter(f"{ns}Item"):
            pos = self._parse_gaeb_item(item, ns)
            if pos:
                positionen.append(pos)

        self.logger.info("GAEB-XML: %d Positionen extrahiert", len(positionen))
        return positionen

    def _parse_gaeb_item(self, item, ns: str) -> dict[str, Any] | None:
        """Parst ein einzelnes GAEB <Item> Element."""

        def find_text(element, tag: str, default: str = "") -> str:
            el = element.find(f".//{ns}{tag}")
            if el is not None and el.text:
                return el.text.strip()
            return default

        # Positionsnummer (RNoPart)
        pos_nr = find_text(item, "RNoPart")
        if not pos_nr:
            # Alternativ: ID-Attribut
            pos_nr = item.get("ID", "")

        if not pos_nr:
            return None

        # Alle span-Texte innerhalb eines Elements sammeln
        def collect_spans(element) -> str:
            parts = []
            for span in element.iter(f"{ns}span"):
                if span.text:
                    parts.append(span.text.strip())
            # Fallback: direkter Text
            if not parts and element.text:
                parts.append(element.text.strip())
            return " ".join(parts)

        # Kurztext: OutlineText > OutlTxt > TextOutlTxt > span
        kurztext = ""
        outline = item.find(f".//{ns}OutlineText")
        if outline is not None:
            kurztext = collect_spans(outline)
        if not kurztext:
            kurztext = find_text(item, "Headline")

        # Langtext: CompleteText > DetailTxt > Text > span
        langtext_parts: list[str] = []
        complete = item.find(f".//{ns}CompleteText")
        if complete is not None:
            langtext_parts.append(collect_spans(complete))
        # Auch alle Text-Elemente durchsuchen
        for text_el in item.iter(f"{ns}Text"):
            text = collect_spans(text_el)
            if text and text not in langtext_parts:
                langtext_parts.append(text)
        langtext = " ".join(langtext_parts)

        if not kurztext and langtext:
            kurztext = langtext[:80]

        # Menge
        qty_text = find_text(item, "Qty", "0")
        try:
            menge = float(qty_text.replace(",", "."))
        except ValueError:
            menge = 0.0

        # Einheit
        einheit = find_text(item, "QU", "STK")

        # Einheitspreis (falls vorbelegt)
        up_text = find_text(item, "UP", "0")
        try:
            einheitspreis = float(up_text.replace(",", "."))
        except ValueError:
            einheitspreis = 0.0

        return {
            "pos_nr": pos_nr,
            "kurztext": kurztext,
            "langtext": langtext,
            "menge": menge,
            "einheit": self._normalize_einheit(einheit),
            "einheitspreis": einheitspreis,
            "gesamtpreis": 0.0,
            "material": "",
            "ist_lackierung": False,
            "ist_fremdleistung": False,
            "sonderanforderungen": [],
            "platten_anzahl": 0,
            "kantenlaenge_lfm": 0,
            "schnittanzahl": 0,
            "bohrungen_anzahl": 0,
        }

    # ------------------------------------------------------------------
    # GAEB D83 Parser (aelteres Textformat)
    # ------------------------------------------------------------------

    def _parse_gaeb_d83(self, pfad: Path) -> list[dict[str, Any]]:
        """Parst GAEB D83 Dateien (Textformat mit festen Satzarten).

        D83 Satzarten:
        - Satzart 25: Positionsdaten (OZ, Menge, Einheit)
        - Satzart 26: Kurztext
        - Satzart 27: Langtext
        """
        self.logger.info("GAEB-D83 parsen: %s", pfad)

        # D83 kann verschiedene Encodings haben
        for encoding in ("cp1252", "latin-1", "utf-8"):
            try:
                content = pfad.read_text(encoding=encoding)
                break
            except (UnicodeDecodeError, UnicodeError):
                continue
        else:
            content = pfad.read_bytes().decode("cp1252", errors="replace")

        lines = content.splitlines()
        positionen: list[dict[str, Any]] = []
        current_pos: dict[str, Any] | None = None

        for line in lines:
            if len(line) < 4:
                continue

            # Satzart steht typisch in Spalte 0-1 oder ist erkennbar am Format
            satzart = line[:2].strip()

            if satzart == "25" or (len(line) > 20 and self._is_position_line(line)):
                # Vorherige Position speichern
                if current_pos:
                    positionen.append(current_pos)

                current_pos = self._parse_d83_position(line)

            elif satzart == "26" and current_pos:
                # Kurztext
                text = line[2:].strip() if len(line) > 2 else ""
                if text:
                    current_pos["kurztext"] = (
                        current_pos.get("kurztext", "") + " " + text
                    ).strip()

            elif satzart == "27" and current_pos:
                # Langtext
                text = line[2:].strip() if len(line) > 2 else ""
                if text:
                    current_pos["langtext"] = (
                        current_pos.get("langtext", "") + " " + text
                    ).strip()

        # Letzte Position
        if current_pos:
            positionen.append(current_pos)

        self.logger.info("GAEB-D83: %d Positionen extrahiert", len(positionen))
        return positionen

    def _is_position_line(self, line: str) -> bool:
        """Heuristik: Ist die Zeile eine Positionszeile?"""
        return bool(re.match(r"^\d{2}\s+[\d.]+\s+", line))

    def _parse_d83_position(self, line: str) -> dict[str, Any]:
        """Parst eine D83-Positionszeile."""
        # Versuche OZ (Ordnungszahl), Menge, Einheit zu extrahieren
        match = re.match(
            r"(?:\d{2}\s+)?([\d.]+)\s+.*?(\d+[.,]?\d*)\s+(\w+)\s*$", line
        )
        if match:
            pos_nr = match.group(1)
            menge_str = match.group(2).replace(",", ".")
            einheit = match.group(3)
            try:
                menge = float(menge_str)
            except ValueError:
                menge = 0.0
        else:
            # Fallback: nur OZ extrahieren
            oz_match = re.match(r"(?:\d{2}\s+)?([\d.]+)", line)
            pos_nr = oz_match.group(1) if oz_match else "?"
            menge = 0.0
            einheit = "STK"

        return {
            "pos_nr": pos_nr,
            "kurztext": "",
            "langtext": "",
            "menge": menge,
            "einheit": self._normalize_einheit(einheit),
            "einheitspreis": 0.0,
            "gesamtpreis": 0.0,
            "material": "",
            "ist_lackierung": False,
            "ist_fremdleistung": False,
            "sonderanforderungen": [],
            "platten_anzahl": 0,
            "kantenlaenge_lfm": 0,
            "schnittanzahl": 0,
            "bohrungen_anzahl": 0,
        }

    # ------------------------------------------------------------------
    # PDF Parser
    # ------------------------------------------------------------------

    def _parse_pdf(self, pfad: Path) -> list[dict[str, Any]]:
        """Parst ein PDF und versucht LV-Positionen zu extrahieren."""
        import pdfplumber

        self.logger.info("PDF parsen: %s", pfad)
        positionen: list[dict[str, Any]] = []

        with pdfplumber.open(str(pfad)) as pdf:
            full_text = ""
            for page in pdf.pages:
                text = page.extract_text()
                if text:
                    full_text += text + "\n"

        if not full_text.strip():
            self.logger.warning("PDF ist leer oder nicht lesbar: %s", pfad)
            return []

        # Positionen aus Text extrahieren (Heuristik)
        positionen = self._extract_positions_from_text(full_text)
        self.logger.info("PDF: %d Positionen extrahiert", len(positionen))
        return positionen

    # ------------------------------------------------------------------
    # Excel Parser
    # ------------------------------------------------------------------

    def _parse_excel(self, pfad: Path) -> list[dict[str, Any]]:
        """Parst eine Excel-Datei mit LV-Positionen."""
        from openpyxl import load_workbook

        self.logger.info("Excel parsen: %s", pfad)

        wb = load_workbook(str(pfad), read_only=True)
        ws = wb.active
        if ws is None:
            return []

        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if not rows:
            return []

        # Header erkennen
        header = [str(h).strip().lower() if h else "" for h in rows[0]]
        positionen: list[dict[str, Any]] = []

        def col(name: str, row_data: tuple) -> str:
            for variant in [name, name.replace("_", " "), name.replace("_", "")]:
                if variant in header:
                    idx = header.index(variant)
                    return str(row_data[idx] or "").strip() if idx < len(row_data) else ""
            return ""

        for row_data in rows[1:]:
            pos_nr = col("pos_nr", row_data) or col("position", row_data) or col("nr", row_data)
            if not pos_nr:
                continue

            menge_str = col("menge", row_data).replace(",", ".")
            try:
                menge = float(menge_str)
            except ValueError:
                menge = 0.0

            positionen.append({
                "pos_nr": pos_nr,
                "kurztext": col("kurztext", row_data) or col("bezeichnung", row_data) or col("text", row_data),
                "langtext": col("langtext", row_data) or col("beschreibung", row_data),
                "menge": menge,
                "einheit": self._normalize_einheit(col("einheit", row_data) or "STK"),
                "einheitspreis": 0.0,
                "gesamtpreis": 0.0,
                "material": col("material", row_data),
                "ist_lackierung": False,
                "ist_fremdleistung": False,
                "sonderanforderungen": [],
                "platten_anzahl": 0,
                "kantenlaenge_lfm": 0,
                "schnittanzahl": 0,
                "bohrungen_anzahl": 0,
            })

        self.logger.info("Excel: %d Positionen extrahiert", len(positionen))
        return positionen

    # ------------------------------------------------------------------
    # Freitext Parser
    # ------------------------------------------------------------------

    def _parse_freitext(self, text: str) -> list[dict[str, Any]]:
        """Interpretiert Freitext-Eingabe und generiert Positionen.

        Einfache Heuristik: Jede Zeile mit erkennbarer Struktur wird zur Position.
        Fuer bessere Ergebnisse: LLM-Router nutzen (agents/llm_router.py).
        """
        self.logger.info("Freitext parsen: %d Zeichen", len(text))
        return self._extract_positions_from_text(text)

    def _extract_positions_from_text(self, text: str) -> list[dict[str, Any]]:
        """Extrahiert Positionen aus beliebigem Text (PDF/Freitext).

        Sucht nach Mustern wie:
        - "01.01.001  Unterschrank 60cm  3 STK"
        - "Pos 1.1: Einbauschrank, Menge: 5"
        """
        positionen: list[dict[str, Any]] = []
        pos_counter = 1

        # Muster: Positionsnummer gefolgt von Text und optional Menge+Einheit
        pattern = re.compile(
            r"(?:Pos\.?\s*)?(\d+(?:\.\d+){0,3})\s+"  # Pos-Nr
            r"(.+?)"                                    # Text
            r"(?:\s+(\d+[.,]?\d*)\s*(STK|St|m2|m²|lfm|PAU|psch|kg|t)\b)?",  # Menge + Einheit
            re.IGNORECASE,
        )

        for line in text.splitlines():
            line = line.strip()
            if not line or len(line) < 5:
                continue

            match = pattern.match(line)
            if match:
                pos_nr = match.group(1)
                kurztext = match.group(2).strip().rstrip(",;:-")
                menge_str = (match.group(3) or "1").replace(",", ".")
                einheit = match.group(4) or "STK"

                try:
                    menge = float(menge_str)
                except ValueError:
                    menge = 1.0

                positionen.append({
                    "pos_nr": pos_nr,
                    "kurztext": kurztext[:120],
                    "langtext": kurztext,
                    "menge": menge,
                    "einheit": self._normalize_einheit(einheit),
                    "einheitspreis": 0.0,
                    "gesamtpreis": 0.0,
                    "material": "",
                    "ist_lackierung": False,
                    "ist_fremdleistung": False,
                    "sonderanforderungen": [],
                    "platten_anzahl": 0,
                    "kantenlaenge_lfm": 0,
                    "schnittanzahl": 0,
                    "bohrungen_anzahl": 0,
                })

        return positionen

    # ------------------------------------------------------------------
    # Hilfsmethoden
    # ------------------------------------------------------------------

    def _normalize_einheit(self, einheit: str) -> str:
        """Normalisiert Einheiten auf Standardwerte."""
        mapping = {
            "st": "STK", "stk": "STK", "stueck": "STK", "stck": "STK",
            "m2": "m2", "m²": "m2", "qm": "m2",
            "lfm": "lfm", "lm": "lfm", "m": "lfm",
            "pau": "PAU", "psch": "PAU", "pauschal": "PAU",
            "kg": "kg", "t": "t",
        }
        return mapping.get(einheit.lower().strip(), einheit.upper())

    def _erkennung_lackierung(
        self, positionen: list[dict[str, Any]]
    ) -> list[dict[str, Any]]:
        """Erkennt Positionen mit Lackieranforderung."""
        lack_keywords = [
            "lackier", "ral ", "ral-", "ncs ", "ncs-",
            "farbbeschicht", "spritzlackier", "pulverbeschicht",
            "hochglanz", "mattlack", "seidenmatt",
        ]

        for pos in positionen:
            texte = f"{pos.get('kurztext', '')} {pos.get('langtext', '')}".lower()
            if any(kw in texte for kw in lack_keywords):
                pos["ist_lackierung"] = True
                pos["ist_fremdleistung"] = True
                pos.setdefault("sonderanforderungen", []).append(
                    "Lackierung erkannt -> Fremdleistung"
                )

        return positionen

    def _detect_projekt_typ(
        self, positionen: list[dict[str, Any]], payload: dict
    ) -> str:
        """Bestimmt den Projekttyp."""
        if payload.get("projekt_typ"):
            return payload["projekt_typ"]

        datei = payload.get("datei_pfad", "").lower()
        if any(datei.endswith(s) for s in (".d83", ".x83", ".x84")):
            return "oeffentlich"

        return "standard"
