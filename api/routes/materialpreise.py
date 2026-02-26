"""
API-Routes für Materialpreisliste.

CRUD + CSV/Excel-Import.
"""

from __future__ import annotations

from datetime import datetime

from fastapi import APIRouter, HTTPException, UploadFile, File

from api.database import get_db
from api.models.schemas import MaterialpreisCreate, MaterialpreisResponse

router = APIRouter(prefix="/api/materialpreise", tags=["Materialpreise"])


@router.get("/", response_model=list[MaterialpreisResponse])
async def liste_materialpreise(kategorie: str = "", suche: str = ""):
    """Materialpreise auflisten, optional gefiltert."""
    async with get_db() as db:
        query = "SELECT * FROM materialpreise WHERE gueltig_bis = ''"
        params: list = []

        if kategorie:
            query += " AND kategorie = ?"
            params.append(kategorie)

        if suche:
            query += " AND material_name LIKE ?"
            params.append(f"%{suche}%")

        query += " ORDER BY material_name"
        cursor = await db.execute(query, params)
        return [dict(row) for row in await cursor.fetchall()]


@router.post("/", response_model=MaterialpreisResponse, status_code=201)
async def erstelle_materialpreis(data: MaterialpreisCreate):
    """Neuen Materialpreis anlegen (versioniert: alter Preis wird archiviert)."""
    jetzt = datetime.now().isoformat()

    async with get_db() as db:
        # Alten Preis archivieren (gueltig_bis setzen)
        await db.execute(
            """UPDATE materialpreise SET gueltig_bis = ?
               WHERE material_name = ? AND gueltig_bis = ''""",
            (jetzt, data.material_name),
        )

        cursor = await db.execute(
            """INSERT INTO materialpreise
               (material_name, kategorie, lieferant, artikel_nr, einheit, preis,
                gueltig_ab, notizen)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
            (
                data.material_name, data.kategorie, data.lieferant,
                data.artikel_nr, data.einheit, data.preis,
                jetzt, data.notizen,
            ),
        )
        await db.commit()

        cursor = await db.execute(
            "SELECT * FROM materialpreise WHERE id = ?", (cursor.lastrowid,)
        )
        return dict(await cursor.fetchone())


@router.post("/import")
async def importiere_preisliste(datei: UploadFile = File(...)):
    """Importiert Materialpreise aus CSV oder Excel.

    Erwartete Spalten: material_name, kategorie, lieferant, artikel_nr, einheit, preis
    """
    if not datei.filename:
        raise HTTPException(status_code=400, detail="Kein Dateiname")

    content = await datei.read()
    suffix = datei.filename.rsplit(".", 1)[-1].lower()

    if suffix == "csv":
        count = await _import_csv(content)
    elif suffix in ("xlsx", "xls"):
        count = await _import_excel(content)
    else:
        raise HTTPException(status_code=400, detail=f"Format nicht unterstützt: {suffix}")

    return {"importiert": count, "datei": datei.filename}


async def _import_csv(content: bytes) -> int:
    """Importiert Preise aus CSV."""
    import csv
    import io

    reader = csv.DictReader(io.StringIO(content.decode("utf-8-sig")))
    count = 0
    jetzt = datetime.now().isoformat()

    async with get_db() as db:
        for row in reader:
            name = row.get("material_name", "").strip()
            if not name:
                continue

            preis = float(row.get("preis", "0").replace(",", "."))

            # Alten Preis archivieren
            await db.execute(
                """UPDATE materialpreise SET gueltig_bis = ?
                   WHERE material_name = ? AND gueltig_bis = ''""",
                (jetzt, name),
            )

            await db.execute(
                """INSERT INTO materialpreise
                   (material_name, kategorie, lieferant, artikel_nr, einheit, preis,
                    gueltig_ab)
                   VALUES (?, ?, ?, ?, ?, ?, ?)""",
                (
                    name,
                    row.get("kategorie", "").strip(),
                    row.get("lieferant", "").strip(),
                    row.get("artikel_nr", "").strip(),
                    row.get("einheit", "STK").strip(),
                    preis,
                    jetzt,
                ),
            )
            count += 1

        await db.commit()
    return count


async def _import_excel(content: bytes) -> int:
    """Importiert Preise aus Excel."""
    import io
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True)
    ws = wb.active
    if ws is None:
        return 0

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return 0

    # Erste Zeile als Header
    header = [str(h).strip().lower() if h else "" for h in rows[0]]
    count = 0
    jetzt = datetime.now().isoformat()

    def col(name: str, row_data: tuple) -> str:
        try:
            idx = header.index(name)
            return str(row_data[idx] or "").strip()
        except (ValueError, IndexError):
            return ""

    async with get_db() as db:
        for row_data in rows[1:]:
            name = col("material_name", row_data)
            if not name:
                continue

            preis_str = col("preis", row_data).replace(",", ".")
            try:
                preis = float(preis_str)
            except ValueError:
                continue

            await db.execute(
                """UPDATE materialpreise SET gueltig_bis = ?
                   WHERE material_name = ? AND gueltig_bis = ''""",
                (jetzt, name),
            )

            await db.execute(
                """INSERT INTO materialpreise
                   (material_name, kategorie, lieferant, artikel_nr, einheit, preis,
                    gueltig_ab)
                   VALUES (?, ?, ?, ?, ?, ?, ?)""",
                (
                    name,
                    col("kategorie", row_data),
                    col("lieferant", row_data),
                    col("artikel_nr", row_data),
                    col("einheit", row_data) or "STK",
                    preis,
                    jetzt,
                ),
            )
            count += 1

        await db.commit()

    wb.close()
    return count
